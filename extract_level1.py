# salva come extract_level1.py e esegui: python extract_level1.py
import openpyxl
import re
import csv
from pathlib import Path

INPUT_XLSX = "1°level.xlsx"
OUTPUT_CSV = "level1_extracted.csv"

# pattern per riconoscere i codici (es. AG5, BX6, A73, L-MH1, C-XLH1, 1L2, etc.)
CODE_RE = re.compile(r"^[A-Z0-9\-]{1,10}$")

wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
# se il foglio ha nome esatto
sheet = wb.active

# Trova le colonne dove compaiono le intestazioni "SEVEL" e "GLIWICE"
sevel_col = None
gliwice_col = None
header_row_idx = None

for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str):
            v = val.strip().upper()
            if v == "SEVEL":
                sevel_col = c
                header_row_idx = r
            if v == "GLIWICE":
                gliwice_col = c
                header_row_idx = r
    if sevel_col and gliwice_col:
        break

# fallback: se non trovate, prova a cercare "SEVEL" o "GLIWICE" in tutto il foglio
if not (sevel_col and gliwice_col):
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if isinstance(val, str):
                v = val.strip().upper()
                if v == "SEVEL" and not sevel_col:
                    sevel_col = c
                    header_row_idx = r
                if v == "GLIWICE" and not gliwice_col:
                    gliwice_col = c
                    header_row_idx = r
    # se ancora mancanti, prosegui comunque: sevel_col o gliwice_col possono rimanere None

# Scansione: manteniamo un body_label corrente
results = []
current_body = None

# Definiamo colonne candidate dove possono comparire i codici (spesso molte colonne)
# useremo tutte le colonne, ma ignoreremo le colonne Sevel/Gliwice
code_columns = list(range(1, sheet.max_column + 1))
if sevel_col in code_columns:
    code_columns.remove(sevel_col)
if gliwice_col in code_columns:
    code_columns.remove(gliwice_col)

for r in range(1, sheet.max_row + 1):
    # aggiorna body_label: cerchiamo nella prima colonna (o nelle prime 3) una cella non vuota che sembra un titolo di sezione
    # regola pratica: se la cella in colonna 1 è testo lungo (più di 3 char) e non è codice, la consideriamo body_label
    first_col_val = sheet.cell(row=r, column=1).value
    if isinstance(first_col_val, str):
        s = first_col_val.strip()
        # se sembra un titolo (contiene spazi e lettere) e non è un codice breve
        if len(s) > 2 and not CODE_RE.match(s.replace(" ", "").upper()):
            current_body = s.upper()
    # inoltre, se troviamo in riga una cella che è esattamente un body label noto (PANEL VAN, GLAZED VAN, ecc.), aggiorniamo
    for c in range(1, 6):  # prime 5 colonne per sicurezza
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, str):
            v = val.strip().upper()
            # lista di parole chiave tipiche
            if any(k in v for k in ["PANEL VAN","GLAZED VAN","CHASSIS","DROPSIDE","CABIN","CASSONATO","CHASSIS COWL","MOTOR COWL","SCUDATO","CHASSIS SPECIAL"]):
                current_body = v
                break

    # ora cerchiamo codici nelle colonne candidate
    for c in code_columns:
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        if val is None:
            continue
        # normalizza valore
        if isinstance(val, str):
            token = val.strip()
        else:
            token = str(val).strip()
        # pulizia token: rimuovi spazi e parentesi
        token_clean = token.split()[0] if " " in token else token
        token_clean = token_clean.replace("\n","").replace("\r","").strip()
        # se il token sembra un codice (es. AG5, L-MH1, A73, BX6)
        if CODE_RE.match(token_clean.upper()):
            sigla = token_clean.upper()
            # controlla Sevel/Gliwice sulla stessa riga
            sevel_flag = False
            gliwice_flag = False
            if sevel_col:
                sval = sheet.cell(row=r, column=sevel_col).value
                if sval is not None:
                    sval_s = str(sval).strip()
                    if sval_s != "-" and sval_s != "":
                        sevel_flag = True
            if gliwice_col:
                gval = sheet.cell(row=r, column=gliwice_col).value
                if gval is not None:
                    gval_s = str(gval).strip()
                    if gval_s != "-" and gval_s != "":
                        gliwice_flag = True
            results.append({
                "sigla_base": sigla,
                "body_label": current_body or "",
                "sevel": "1" if sevel_flag else "0",
                "gliwice": "1" if gliwice_flag else "0",
                "notes": ""
            })

# deduplica mantenendo prima occorrenza
seen = set()
unique = []
for r in results:
    key = (r["sigla_base"], r["body_label"])
    if key not in seen:
        seen.add(key)
        unique.append(r)

# scrivi CSV
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["sigla_base","body_label","sevel","gliwice","notes"])
    writer.writeheader()
    for row in unique:
        writer.writerow(row)

print(f"Estrazione completata. {len(unique)} righe scritte in {OUTPUT_CSV}")
